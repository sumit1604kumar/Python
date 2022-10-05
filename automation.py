import openpyxl # external package from pypi website

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file['Sheet1']   # file name in the excel means the TAB
# task 1 to calculate how many products/per supplier
product_per_supplier = {}  # empty dictonaries
total_value_per_supplier = {}
products_under_10_inv = {}
product_list.cell(1, 5).value = 'Patel Title'



for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_number = product_list.cell(product_row, 1).value
    new_cell = product_list.cell(product_row, 5)
# calculation of number of products/supplier
    if supplier_name in product_per_supplier:
        current_num_products = product_per_supplier.get(supplier_name)
        product_per_supplier[supplier_name] = current_num_products + 1

    else:
        product_per_supplier[supplier_name] = 1

#task 2 total value of inventory/supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

#task 2 -print the products which has inventory less than 10
    if  inventory < 10:
        products_under_10_inv[int(product_number)] = int(inventory)
# TASK 4
    new_cell.value = inventory * price

inv_file.save("anju.xlsx")
print(product_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)





    




